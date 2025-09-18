from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
    QTableWidgetItem, QPushButton, QLineEdit, QComboBox,
    QLabel, QMessageBox, QHeaderView, QInputDialog, QDialog,
    QFormLayout, QDialogButtonBox, QGridLayout
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import QIcon
import logging


class ClassManagementDialog(QDialog):
    """班级管理对话框"""

    def __init__(self, db_conn):
        super().__init__()
        self.db_conn = db_conn
        self.logger = logging.getLogger(__name__)
        self.setWindowTitle("班级管理")
        self.setWindowIcon(QIcon('img/icon.png'))
        self.resize(500, 400)
        self.init_ui()
        self.load_classes()

    def init_ui(self):
        layout = QVBoxLayout()

        # 添加班级区域
        add_layout = QHBoxLayout()
        self.class_name_input = QLineEdit()
        self.class_name_input.setPlaceholderText("请输入班级名称")
        add_btn = QPushButton("添加班级")
        add_btn.clicked.connect(self.add_class)
        add_layout.addWidget(self.class_name_input)
        add_layout.addWidget(add_btn)
        layout.addLayout(add_layout)

        # 班级表格
        self.class_table = QTableWidget()
        self.class_table.setColumnCount(2)
        self.class_table.setHorizontalHeaderLabels(["班级ID", "班级名称"])
        self.class_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.class_table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.class_table)

        # 操作按钮
        btn_layout = QHBoxLayout()
        delete_btn = QPushButton("删除选中班级")
        delete_btn.clicked.connect(self.delete_class)
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(delete_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def load_classes(self):
        """加载班级列表"""
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT class_id, class_name FROM classes ORDER BY class_id")
            classes = cursor.fetchall()

            self.class_table.setRowCount(len(classes))
            for row, class_data in enumerate(classes):
                self.class_table.setItem(row, 0, QTableWidgetItem(str(class_data[0])))
                self.class_table.setItem(row, 1, QTableWidgetItem(class_data[1]))
        except Exception as e:
            self.logger.error("加载班级列表错误: %s", str(e))
            QMessageBox.critical(self, "错误", f"加载班级列表失败: {str(e)}")

    def add_class(self):
        """添加班级"""
        class_name = self.class_name_input.text().strip()
        if not class_name:
            QMessageBox.warning(self, "提示", "班级名称不能为空")
            return

        try:
            cursor = self.db_conn.cursor()
            cursor.execute("INSERT INTO classes (class_name) VALUES (?)", (class_name,))
            self.db_conn.commit()
            self.class_name_input.clear()
            self.load_classes()
            self.logger.info("添加班级: %s", class_name)
            QMessageBox.information(self, "成功", f"班级 '{class_name}' 添加成功")
        except Exception as e:
            self.logger.error("添加班级错误: %s", str(e))
            QMessageBox.critical(self, "错误", f"添加班级失败: {str(e)}")

    def delete_class(self):
        """删除班级"""
        selected = self.class_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "提示", "请先选择要删除的班级")
            return

        class_id = self.class_table.item(selected[0].row(), 0).text()
        class_name = self.class_table.item(selected[0].row(), 1).text()

        # 检查班级中是否有学生
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM students WHERE class_id = ?", (class_id,))
            student_count = cursor.fetchone()[0]

            if student_count > 0:
                QMessageBox.warning(self, "提示",
                                    f"班级 '{class_name}' 中还有 {student_count} 名学生，无法删除")
                return
        except Exception as e:
            self.logger.error("检查班级学生错误: %s", str(e))
            QMessageBox.critical(self, "错误", f"检查班级学生失败: {str(e)}")
            return

        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除班级 '{class_name}' 吗？此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("DELETE FROM classes WHERE class_id = ?", (class_id,))
                self.db_conn.commit()
                self.load_classes()
                self.logger.info("删除班级: %s", class_name)
                QMessageBox.information(self, "成功", f"班级 '{class_name}' 删除成功")
            except Exception as e:
                self.logger.error("删除班级错误: %s", str(e))
                QMessageBox.critical(self, "错误", f"删除班级失败: {str(e)}")


class StudentEditDialog(QDialog):
    """学生信息编辑对话框（简化版）"""

    def __init__(self, db_conn, student_data=None):
        super().__init__()
        self.db_conn = db_conn
        self.student_data = student_data
        self.setWindowTitle("编辑学生信息" if student_data else "添加学生")
        self.setWindowIcon(QIcon('img/icon.png'))
        self.resize(300, 150)
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout()

        # 学号
        self.student_id_input = QLineEdit()
        self.student_id_input.setPlaceholderText("请输入学号")
        layout.addRow("学号:", self.student_id_input)

        # 姓名
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("请输入姓名")
        layout.addRow("姓名:", self.name_input)

        # 班级选择
        self.class_combo = QComboBox()
        self.load_classes()
        layout.addRow("班级:", self.class_combo)

        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

        self.setLayout(layout)

        # 如果是编辑模式，填充数据
        if self.student_data:
            self.fill_data()

    def load_classes(self):
        """加载班级列表"""
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT class_id, class_name FROM classes ORDER BY class_name")
            classes = cursor.fetchall()

            self.class_combo.addItem("未分配班级", None)
            for class_id, class_name in classes:
                self.class_combo.addItem(class_name, class_id)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载班级列表失败: {str(e)}")

    def fill_data(self):
        """填充学生数据"""
        if not self.student_data:
            return

        self.student_id_input.setText(self.student_data.get('student_id', ''))
        self.name_input.setText(self.student_data.get('name', ''))

        # 设置班级
        class_id = self.student_data.get('class_id')
        if class_id:
            for i in range(self.class_combo.count()):
                if self.class_combo.itemData(i) == class_id:
                    self.class_combo.setCurrentIndex(i)
                    break

    def get_data(self):
        """获取表单数据"""
        return {
            'student_id': self.student_id_input.text().strip(),
            'name': self.name_input.text().strip(),
            'class_id': self.class_combo.currentData()
        }


class StudentManagementWindow(QWidget):
    """学生管理窗口（简化版）"""

    def __init__(self, db_conn):
        super().__init__()
        self.db_conn = db_conn
        self.logger = logging.getLogger(__name__)

        self.setWindowTitle("学生管理")
        self.setWindowIcon(QIcon('img/icon.png'))
        self.resize(800, 600)
        self.init_ui()
        self.load_students()

    def init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout()

        # 搜索和过滤区域
        search_layout = QHBoxLayout()

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索学生姓名或学号...")
        self.search_input.textChanged.connect(self.load_students)

        self.class_filter = QComboBox()
        self.class_filter.addItem("所有班级", None)
        self.load_classes()
        self.class_filter.currentIndexChanged.connect(self.load_students)

        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.class_filter)
        layout.addLayout(search_layout)

        # 操作按钮区域
        btn_layout = QGridLayout()
        btn_layout.setSpacing(5)

        # 第一行按钮
        add_btn = QPushButton("添加学生")
        add_btn.clicked.connect(self.add_student)

        edit_btn = QPushButton("编辑学生")
        edit_btn.clicked.connect(self.edit_student)

        delete_btn = QPushButton("删除学生")
        delete_btn.clicked.connect(self.delete_student)

        # 第二行按钮
        class_btn = QPushButton("班级管理")
        class_btn.clicked.connect(self.manage_classes)

        import_btn = QPushButton("导入学生")
        import_btn.clicked.connect(self.import_students)

        export_btn = QPushButton("导出学生")
        export_btn.clicked.connect(self.export_students)

        btn_layout.addWidget(add_btn, 0, 0)
        btn_layout.addWidget(edit_btn, 0, 1)
        btn_layout.addWidget(delete_btn, 0, 2)
        btn_layout.addWidget(class_btn, 1, 0)
        btn_layout.addWidget(import_btn, 1, 1)
        btn_layout.addWidget(export_btn, 1, 2)

        layout.addLayout(btn_layout)

        # 学生表格
        self.student_table = QTableWidget()
        self.student_table.setColumnCount(3)
        self.student_table.setHorizontalHeaderLabels(["学号", "姓名", "班级"])
        self.student_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.student_table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.student_table)

        self.setLayout(layout)

    def _check_db_connection(self):
        """检查数据库连接是否有效"""
        if self.db_conn is None:
            QMessageBox.critical(self, "错误", "数据库连接无效")
            return False
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT 1")
            cursor.fetchone()
            return True
        except Exception as e:
            QMessageBox.critical(self, "错误", f"数据库连接已关闭: {str(e)}")
            return False

    def load_classes(self):
        """加载班级列表"""
        if not self._check_db_connection():
            return

        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT class_id, class_name FROM classes ORDER BY class_name")
            classes = cursor.fetchall()

            # 保存当前选中的班级
            current_class = self.class_filter.currentData()

            self.class_filter.clear()
            self.class_filter.addItem("所有班级", None)

            for class_id, class_name in classes:
                self.class_filter.addItem(class_name, class_id)

            # 恢复之前选中的班级
            if current_class:
                for i in range(self.class_filter.count()):
                    if self.class_filter.itemData(i) == current_class:
                        self.class_filter.setCurrentIndex(i)
                        break

        except Exception as e:
            self.logger.error("加载班级列表错误: %s", str(e))
            QMessageBox.critical(self, "错误", f"加载班级列表失败: {str(e)}")

    def load_students(self):
        """加载学生列表"""
        if not self._check_db_connection():
            return

        search_text = self.search_input.text().strip()
        class_id = self.class_filter.currentData()

        try:
            cursor = self.db_conn.cursor()

            query = """
                    SELECT s.student_id, s.name, c.class_name
                    FROM students s
                             LEFT JOIN classes c ON s.class_id = c.class_id
                    WHERE 1 = 1 \
                    """
            params = []

            if search_text:
                query += " AND (s.name LIKE ? OR s.student_id LIKE ?)"
                params.extend([f"%{search_text}%", f"%{search_text}%"])

            if class_id:
                query += " AND s.class_id = ?"
                params.append(class_id)

            query += " ORDER BY s.student_id"

            cursor.execute(query, params)
            students = cursor.fetchall()

            self.student_table.setRowCount(len(students))
            for row, student in enumerate(students):
                for col in range(3):
                    value = str(student[col]) if student[col] is not None else ""
                    self.student_table.setItem(row, col, QTableWidgetItem(value))

        except Exception as e:
            self.logger.error("加载学生列表错误: %s", str(e))
            QMessageBox.critical(self, "错误", f"加载学生列表失败: {str(e)}")

    def manage_classes(self):
        """班级管理"""
        dialog = ClassManagementDialog(self.db_conn)
        dialog.exec_()
        # 刷新班级列表
        self.load_classes()
        # 刷新学生列表
        self.load_students()

    def add_student(self):
        """添加学生"""
        dialog = StudentEditDialog(self.db_conn)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()

            if not data['student_id']:
                QMessageBox.warning(self, "提示", "学号不能为空")
                return

            if not data['name']:
                QMessageBox.warning(self, "提示", "姓名不能为空")
                return

            try:
                cursor = self.db_conn.cursor()
                cursor.execute("""
                               INSERT INTO students (student_id, name, class_id)
                               VALUES (?, ?, ?)
                               """, (data['student_id'], data['name'], data['class_id']))
                self.db_conn.commit()
                self.load_students()
                self.logger.info("添加学生: %s", data['student_id'])
                QMessageBox.information(self, "成功", f"学生 {data['name']} 添加成功")
            except Exception as e:
                self.logger.error("添加学生错误: %s", str(e))
                QMessageBox.critical(self, "错误", f"添加学生失败: {str(e)}")

    def edit_student(self):
        """编辑学生信息"""
        selected = self.student_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "提示", "请先选择要编辑的学生")
            return

        student_id = self.student_table.item(selected[0].row(), 0).text()

        try:
            cursor = self.db_conn.cursor()
            cursor.execute("""
                           SELECT student_id, name, class_id
                           FROM students
                           WHERE student_id = ?
                           """, (student_id,))
            student_data = cursor.fetchone()

            if student_data:
                data_dict = {
                    'student_id': student_data[0],
                    'name': student_data[1],
                    'class_id': student_data[2]
                }

                dialog = StudentEditDialog(self.db_conn, data_dict)
                if dialog.exec_() == QDialog.Accepted:
                    new_data = dialog.get_data()

                    cursor.execute("""
                                   UPDATE students
                                   SET name     = ?,
                                       class_id = ?
                                   WHERE student_id = ?
                                   """, (new_data['name'], new_data['class_id'], student_id))
                    self.db_conn.commit()
                    self.load_students()
                    self.logger.info("更新学生: %s", student_id)
                    QMessageBox.information(self, "成功", f"学生信息更新成功")

        except Exception as e:
            self.logger.error("编辑学生错误: %s", str(e))
            QMessageBox.critical(self, "错误", f"编辑学生失败: {str(e)}")

    def delete_student(self):
        """删除学生"""
        selected = self.student_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "提示", "请先选择要删除的学生")
            return

        student_id = self.student_table.item(selected[0].row(), 0).text()
        student_name = self.student_table.item(selected[0].row(), 1).text()

        reply = QMessageBox.question(
            self, "确认删除",
            f"确定要删除学生 {student_name} ({student_id}) 吗？此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                cursor = self.db_conn.cursor()
                cursor.execute("DELETE FROM students WHERE student_id = ?", (student_id,))
                self.db_conn.commit()
                self.load_students()
                self.logger.info("删除学生: %s", student_id)
                QMessageBox.information(self, "成功", f"学生 {student_name} 删除成功")
            except Exception as e:
                self.logger.error("删除学生错误: %s", str(e))
                QMessageBox.critical(self, "错误", f"删除学生失败: {str(e)}")

    def import_students(self):
        """从Excel导入学生数据"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel文件",
            "",
            "Excel文件 (*.xlsx *.xls)"
        )

        if not file_path:
            return  # 用户取消了选择

        try:
            # 加载Excel文件
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            students = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过标题行
                if row and row[0] and row[1]:  # 确保学号和姓名不为空
                    students.append((str(row[0]).strip(), str(row[1]).strip()))

            if not students:
                QMessageBox.warning(self, "提示", "Excel中没有找到有效学生数据")
                return

            # 选择班级
            class_id = self._select_class_for_import()
            if class_id is None:  # 用户取消了选择
                return

            # 确认导入
            reply = QMessageBox.question(
                self, "确认导入",
                f"即将导入 {len(students)} 名学生到选定的班级\n学号: {students[0][0]} 到 {students[-1][0]}\n确认继续吗?",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                self._save_imported_students(students, class_id)

        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入Excel失败: {str(e)}")
            self.logger.error("导入Excel错误: %s", str(e))

    def _select_class_for_import(self):
        """选择导入学生的班级"""
        dialog = QDialog(self)
        dialog.setWindowTitle("选择班级")
        dialog.resize(300, 150)

        layout = QVBoxLayout()

        label = QLabel("请选择要导入的班级:")
        layout.addWidget(label)

        class_combo = QComboBox()
        try:
            cursor = self.db_conn.cursor()
            cursor.execute("SELECT class_id, class_name FROM classes ORDER BY class_name")
            classes = cursor.fetchall()

            class_combo.addItem("未分配班级", None)
            for class_id, class_name in classes:
                class_combo.addItem(class_name, class_id)
        except Exception as e:
            self.logger.error("加载班级列表错误: %s", str(e))
            QMessageBox.critical(self, "错误", f"加载班级列表失败: {str(e)}")
            return None

        layout.addWidget(class_combo)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        dialog.setLayout(layout)

        if dialog.exec_() == QDialog.Accepted:
            return class_combo.currentData()
        return None

    def _save_imported_students(self, students, class_id):
        """将导入的学生保存到数据库"""
        cursor = self.db_conn.cursor()
        success_count = 0
        duplicate_count = 0

        try:
            for student_id, name in students:
                # 检查学号是否已存在
                cursor.execute("SELECT 1 FROM students WHERE student_id = ?", (student_id,))
                if cursor.fetchone():
                    duplicate_count += 1
                    self.logger.warning("学号已存在，跳过: %s", student_id)
                    continue

                # 插入新学生
                cursor.execute("""
                               INSERT INTO students (student_id, name, class_id)
                               VALUES (?, ?, ?)
                               """, (student_id, name, class_id))
                success_count += 1

            self.db_conn.commit()
            self.load_students()  # 刷新表格

            message = f"成功导入 {success_count} 名学生"
            if duplicate_count > 0:
                message += f"\n跳过 {duplicate_count} 名已存在的学生"

            QMessageBox.information(self, "导入完成", message)
            self.logger.info("从Excel导入 %d 名学生", success_count)

        except Exception as e:
            self.db_conn.rollback()
            QMessageBox.critical(self, "错误", f"保存学生数据失败: {str(e)}")
            self.logger.error("保存导入学生错误: %s", str(e))

    def export_students(self):
        """导出学生数据"""
        QMessageBox.information(self, "提示", "导出学生功能开发中")