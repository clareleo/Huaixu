# score_editor_window.py

import os
import re
from typing import List, Dict, Optional
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QFileDialog, QMessageBox, QComboBox, QLineEdit,
    QHeaderView, QAbstractItemView, QGroupBox, QInputDialog
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIntValidator


class ScoreEditorWindow(QWidget):
    def __init__(self, db_conn, parent=None):
        super().__init__(parent)
        self.db_conn = db_conn
        self.setWindowTitle("成绩编辑器 - 支持4+X模板")
        self.resize(1200, 800)

        # 内部数据结构
        self.students: List[Dict] = []          # 学生列表
        self.fixed_columns = ["学号", "姓名"]     # 固定列
        self.process_4_items = []                # 过程评价“4”项（考勤、课堂表现等）
        self.process_x_items = []                # 过程评价“X”项（动态）
        self.final_exam_col = "期末考试"
        self.metadata = {"专业部": "", "班级": "", "学科": ""}

        self.init_ui()
        self.setup_db()

    def init_ui(self):
        layout = QVBoxLayout()

        # === 顶部信息栏 ===
        info_layout = QHBoxLayout()
        self.dept_label = QLabel("专业部：--")
        self.class_label = QLabel("班级：--")
        self.subject_label = QLabel("学科：--")
        info_layout.addWidget(self.dept_label)
        info_layout.addWidget(self.class_label)
        info_layout.addWidget(self.subject_label)
        info_layout.addStretch()
        layout.addLayout(info_layout)

        # === 控制区 ===
        control_group = QGroupBox("操作")
        control_layout = QHBoxLayout()
        self.load_btn = QPushButton("📂 导入Excel模板")
        self.add_x_btn = QPushButton("+ 添加X项")
        self.save_btn = QPushButton("💾 保存到数据库")
        self.export_btn = QPushButton("📤 导出Excel")

        self.load_btn.clicked.connect(self.load_excel_template)
        self.add_x_btn.clicked.connect(self.add_x_item)
        self.save_btn.clicked.connect(self.save_to_db)
        self.export_btn.clicked.connect(self.export_to_excel)

        control_layout.addWidget(self.load_btn)
        control_layout.addWidget(self.add_x_btn)
        control_layout.addWidget(self.save_btn)
        control_layout.addWidget(self.export_btn)
        control_group.setLayout(control_layout)
        layout.addWidget(control_group)

        # === 成绩表格 ===
        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        layout.addWidget(self.table)

        self.setLayout(layout)

    def setup_db(self):
        """确保成绩表存在"""
        cursor = self.db_conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS scores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                department TEXT,
                class_name TEXT,
                subject TEXT,
                student_id TEXT,
                student_name TEXT,
                process_4_json TEXT,   -- JSON字符串存储4项
                process_x_json TEXT,   -- JSON字符串存储X项
                final_exam REAL,
                total_score REAL,
                credit REAL DEFAULT 1.0
            )
        """)
        self.db_conn.commit()

    def load_excel_template(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择成绩模板Excel文件",
            "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # 自动识别元数据（从合并单元格或特定位置）
            self.parse_metadata(ws)

            # 解析表头（假设第5行为字段行）
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=5, column=col).value
                if cell:
                    headers.append(str(cell).strip())
                else:
                    headers.append(f"Unnamed_{col}")

            # 分离字段类型
            self.process_4_items = []
            self.process_x_items = []
            other_cols = []

            for h in headers:
                if h in ["学号", "姓名"]:
                    continue
                elif "考勤" in h or "课堂表现" in h or "作业" in h or "测验" in h:
                    self.process_4_items.append(h)
                elif "笔记" in h or ("X" in h and "过程" not in h):
                    self.process_x_items.append(h)
                elif "期末" in h:
                    self.final_exam_col = h
                else:
                    other_cols.append(h)

            # 读取学生数据（从第6行开始）
            self.students = []
            for row in range(6, ws.max_row + 1):
                if not ws.cell(row=row, column=1).value:
                    break  # 空行终止
                student = {}
                for i, header in enumerate(headers):
                    val = ws.cell(row=row, column=i+1).value
                    student[header] = str(val) if val is not None else ""
                self.students.append(student)

            self.refresh_table()
            self.update_metadata_labels()

        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载Excel失败：{str(e)}")

    def parse_metadata(self, ws):
        """尝试从Excel中提取专业部、班级、学科"""
        # 示例：A1 单元格可能是 "信息技术专业部"
        # A2 可能是 "2022级计算机应用1班"
        # A3 可能是 "课程：Python程序设计"

        dept = str(ws.cell(1, 1).value or "").strip()
        cls = str(ws.cell(2, 1).value or "").strip()
        subj_line = str(ws.cell(3, 1).value or "").strip()

        # 提取学科（如“课程：数学” → “数学”）
        subject = re.sub(r".*[:：]\s*", "", subj_line) if "课程" in subj_line else subj_line

        self.metadata = {
            "专业部": dept if dept and dept != "None" else "--",
            "班级": cls if cls and cls != "None" else "--",
            "学科": subject if subject and subject != "None" else "--"
        }

    def update_metadata_labels(self):
        self.dept_label.setText(f"专业部：{self.metadata['专业部']}")
        self.class_label.setText(f"班级：{self.metadata['班级']}")
        self.subject_label.setText(f"学科：{self.metadata['学科']}")

    def refresh_table(self):
        all_columns = (
            self.fixed_columns +
            self.process_4_items +
            self.process_x_items +
            [self.final_exam_col, "总分"]
        )

        self.table.setRowCount(len(self.students))
        self.table.setColumnCount(len(all_columns))
        self.table.setHorizontalHeaderLabels(all_columns)

        # 设置分数列只允许输入数字
        validator = QIntValidator(0, 100)

        for row_idx, student in enumerate(self.students):
            for col_idx, col in enumerate(all_columns):
                item = QTableWidgetItem()
                if col in ["学号", "姓名"]:
                    item.setText(student.get(col, ""))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                elif col == "总分":
                    item.setText(self.calculate_total(student))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                else:
                    val = student.get(col, "")
                    item.setText(str(val) if val else "")
                    item.setData(Qt.UserRole, col)  # 用于校验
                self.table.setItem(row_idx, col_idx, item)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

    def calculate_total(self, student: dict) -> str:
        try:
            # 计算过程评价50%
            p4_sum = sum(float(student.get(k, 0) or 0) for k in self.process_4_items)
            px_sum = sum(float(student.get(k, 0) or 0) for k in self.process_x_items)
            process_total = p4_sum + px_sum  # 理论上应 ≤ 50，但按实际值算
            final = float(student.get(self.final_exam_col, 0) or 0)
            total = round((process_total + final) / 2, 1)  # 假设过程50%+期末50%
            return f"{total:.1f}"
        except:
            return "0.0"

    def add_x_item(self):
        new_name, ok = QInputDialog.getText(self, "新增X项", "请输入新评价项名称（如：项目实践）：")
        if ok and new_name.strip():
            new_name = new_name.strip()
            if new_name not in self.process_x_items:
                self.process_x_items.append(new_name)
                # 为所有学生初始化空值
                for s in self.students:
                    s[new_name] = ""
                self.refresh_table()
            else:
                QMessageBox.warning(self, "提示", "该项已存在！")

    def save_to_db(self):
        if not self.students:
            QMessageBox.warning(self, "警告", "无数据可保存！")
            return

        cursor = self.db_conn.cursor()
        import json

        for s in self.students:
            # 校验分数
            for col in self.process_4_items + self.process_x_items + [self.final_exam_col]:
                val = s.get(col, "")
                if val == "":
                    s[col] = 0
                else:
                    try:
                        num = float(val)
                        if not (0 <= num <= 100):
                            raise ValueError
                        s[col] = num
                    except:
                        QMessageBox.warning(self, "数据错误", f"学生 {s['姓名']} 的 {col} 分数无效（应为0-100）")
                        return

            total = float(self.calculate_total(s))

            cursor.execute("""
                INSERT INTO scores 
                (department, class_name, subject, student_id, student_name,
                 process_4_json, process_x_json, final_exam, total_score)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                self.metadata["专业部"],
                self.metadata["班级"],
                self.metadata["学科"],
                s["学号"],
                s["姓名"],
                json.dumps({k: s[k] for k in self.process_4_items}),
                json.dumps({k: s[k] for k in self.process_x_items}),
                s[self.final_exam_col],
                total
            ))

        self.db_conn.commit()
        QMessageBox.information(self, "成功", f"已保存 {len(self.students)} 条成绩记录！")

    def export_to_excel(self):
        QMessageBox.information(self, "提示", "导出功能待实现（保留原模板格式）")
        # 此处可调用 openpyxl 写回模板，略

    # ======================
    # 表格编辑时实时校验（可选增强）
    # ======================
    def closeEvent(self, event):
        # 可添加未保存提醒
        event.accept()




